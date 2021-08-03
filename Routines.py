def matrixvuller ( filenaam,  kolommen, beginrij = 2, beginkolom = 2,   sheetnaam = 'Sheet1') :
    import openpyxl
    wb2=openpyxl.load_workbook(filenaam)
    sheet2=wb2[sheetnaam]
    matrix = []
    aantal_rijen = sheet2.max_row

    for r in range (beginrij, aantal_rijen + 1):
        matrix.append ([])
        for k in range (beginkolom, kolommen + beginkolom):
            matrixcel = sheet2.cell(row=r, column=k)
            Waarde =  matrixcel.value
            matrix [r-beginrij] .append (Waarde)
    return matrix

def kolomkopvuller (filenaam, sheetnaam) :
    import openpyxl
    wb2=openpyxl.load_workbook(filenaam)
    sheet2=wb2[sheetnaam]
    kolomkoppen = []

    for k in range ( 2, sheet2.max_column + 1):
            koppencel = sheet2.cell(row=1, column=k)
            Koptekst = koppencel.value
            kolomkoppen .append (Koptekst)
    return kolomkoppen


def zoekrijnummer (i, kolom):
    for j in range ( len ( kolom ) ):
        if kolom[j] == i:
            rijnummer = j
            break
    return rijnummer

def zoekSEGkolom(kolomkoppen, tekst):
    for k in range (0,len(kolomkoppen)):
        if kolomkoppen[k] == tekst :
            zoekkolom = k
            break
    return zoekkolom

def matrixvolnullen (rijen,kolommen) :
    matrix = []
    for r in range (0,rijen):
        matrix.append ([])
        for k in range (0,kolommen):
            matrix[r].append (0)
    return matrix

def lijstvolnullen (lengte) :
    lijst = []
    for r in range (0,lengte):
        lijst.append (0)
    return lijst

def xlswegschrijven (matrix, filenaam):
    import xlsxwriter
    workbook = xlsxwriter.Workbook ( filenaam + '.xlsx' )
    worksheet = workbook.add_worksheet ( )
    for r in range ( 0, 1425 ):
        worksheet.write_row ( r + 1, 1, matrix[r] )
    workbook.close ( )

def transponeren (matrix, breedte, hoogte):
    uitmatrix = []
    for i in range (breedte):
        uitmatrix.append([])
        for j in range (hoogte):
            uitmatrix[i].append(matrix[j][i])
    return (uitmatrix)


def xlswegschrijven_totalen (matrix, header, getallenlijst, filenaam, aantal_zones=1425):
    import xlsxwriter
    import Routines
    transmatrix = Routines.transponeren (matrix, aantal_zones, len(header)-1)
    workbook = xlsxwriter.Workbook ( filenaam + '.xlsx' )
    worksheet = workbook.add_worksheet ( )
    worksheet.write_row (0,0,header)
    worksheet.write_column (1,0,getallenlijst)
    for r in range ( 0, 1425 ):
        worksheet.write_row ( r + 1, 1, transmatrix[r] )
    workbook.close ( )

def matrixen_maken_voor_Excel_totalen (Totalendirectory, groep):
    import Routines
    vervoerscombis = ['Fiets', 'EFiets', 'Auto', 'OV', 'Fiets_Auto', 'Fiets_OV', 'EFiets_Auto', 'Efiets_OV', 'Auto_OV',
                      'Fiets_Auto_OV', 'Efiets_Auto_OV']
    Matrix = []

    for vvcombis in vervoerscombis:
        if not groep is None:
            Matrix.append(Routines.csvlezen (Totalendirectory + vvcombis + '_' + str(groep)))
        else:
            Matrix.append ( Routines.csvlezen ( Totalendirectory + vvcombis  ) )
    return Matrix

def getallenlijst_maken (aantal_getallen):
    Lijst = []
    for i in range (aantal_getallen):
        Lijst.append(i+1)
    return Lijst




def csvlezen (filenaam, aantal_lege_regels=0):
    import Routines
    import csv
    matrix = []
    filenaam2=filenaam + '.csv'
    with open ( filenaam2, 'r' ) as csvfile:
        reader = csv.reader ( csvfile )
        for i in range (aantal_lege_regels):
            next(reader)
        for row in reader:
            matrix.append ( row )
    lengte = len ( matrix[0] )
    with open ( filenaam2, 'r' ) as csvfile:
        reader = csv.reader ( csvfile )
        row_count = sum ( 1 for row in reader )
    uitmatrix = []
    tussenmatrix = []
    if row_count == 1:
        tussenmatrix.append ( matrix[0] )
        for elem in tussenmatrix[0]:
            uitmatrix.append ( float ( elem ) )
    else:
        for r in range ( 0, row_count - aantal_lege_regels ):
            tussenmatrix.append (matrix[r])
            uitmatrix.append ([])
            for elem in tussenmatrix[r]:
                uitmatrix[r].append ( float(elem) )
    return uitmatrix

def Omnitrans_csv_inlezen (filenaam, aantal_zones=1425, aantal_lege_regels=4):
    import Routines
    import csv
    lijst = []
    skimmatrix = []
    filenaam2=filenaam + '.csv'
    with open ( filenaam2, 'r' ) as csvfile:
        reader = csv.reader ( csvfile )
        row_count = sum ( 1 for row in reader )
    with open ( filenaam2, 'r' ) as csvfile:
        reader = csv.reader ( csvfile )
        for i in range (aantal_lege_regels):
            next(reader)
        for row in reader:
            lijst.append(row)
    for i in range (aantal_zones):
        skimmatrix.append([])
        for j in range (aantal_zones):
            r = i*1425+j
            skimmatrix[i].append(float(lijst[r][2]))
    return skimmatrix

def csvwegschrijven (matrix, filenaam, soort = "matrix"):
    import csv
    f = open ( filenaam + '.csv', 'w', newline='' )
    with f:
        writer = csv.writer ( f )
        if soort == "matrix":
            writer.writerows ( matrix )
        else:
            writer.writerow ( matrix )


def laad_constanten_reistijdvervalscurve (Gewichtendirectory,  vvwijze, groep, motief ):
    import openpyxl
    zoekkolom = {'werk':2,'winkeldagelijks':6, 'winkelnietdagelijks':10, 'onderwijs':14, 'zorg':18, 'overig':22}
    optelwaarde = {'Auto':0, 'Fiets':1, 'EFiets':2, 'OV':3}
    filenaam2 = Gewichtendirectory + 'Constanten_reistijdvervalscurve.xlsx'
    wb=openpyxl.load_workbook(filenaam2)
    sheet=wb['Alphas']
    sheet2=wb['Omegas']
    sheet3=wb['Weging']
    alphacel = sheet.cell(row=groep + 3, column= zoekkolom.get(motief) + optelwaarde.get(vvwijze))
    alpha = alphacel.value
    omegacel = sheet2.cell(row=groep + 3, column= zoekkolom.get(motief) + optelwaarde.get(vvwijze))
    omega = omegacel.value
    return alpha, omega

def laad_alle_constanten_reistijdvervalcurve (Gewichtendirectory, motief, aantal_groepen = 20):
    import Routines
    vervoerwijze=['Auto', 'Fiets', 'EFiets', 'OV']

    for vvwijze in vervoerwijze:
        Alphas = []
        Omegas = []
        for groep in range (0, aantal_groepen + 1):
            Alphasnaam = 'Alphas_'+ vvwijze
            Omegasnaam = 'Omegas_' + vvwijze
            Alphaomega = Routines.laad_constanten_reistijdvervalscurve(Gewichtendirectory,vvwijze,groep, motief)
            Alphas.append(Alphaomega[0])
            Omegas.append(Alphaomega[1])
        Alphafilenaam = Gewichtendirectory + Alphasnaam
        Omegafilenaam = Gewichtendirectory + Omegasnaam
        Routines.csvwegschrijven(Alphas,Alphafilenaam, soort = "lijst")
        Routines.csvwegschrijven(Omegas,Omegafilenaam, soort = "lijst")

def maak_totale_buurten_file (Invoerexcel, aantal_zones = 1425):
    import Routines
    Zonelijst_met_buurtsamenstelling = Routines.matrixvuller ( Invoerexcel, 1, beginrij=3, beginkolom=1,
                                                               sheetnaam='Excl_studenten' )
    Zonelijst = [sub[0] for sub in Zonelijst_met_buurtsamenstelling]
    Buurtsamenstelling = Routines.matrixvuller (Invoerexcel, 20, beginrij=3, beginkolom=2,
                                                 sheetnaam='Excl_studenten' )
    Zonelijstnummer = 0
    Samenstelling_alle_zones = []
    for i in range ( aantal_zones ):
        Samenstelling_alle_zones.append ( [] )
        if i + 1 in Zonelijst:
            for k in range ( 0, 20 ):
                Samenstelling_alle_zones[i].append ( Buurtsamenstelling[Zonelijstnummer][k] )
            Zonelijstnummer = Zonelijstnummer + 1
        else:
            for k in range ( 0, 20 ):
                Samenstelling_alle_zones[i].append ( 0 )
    return Samenstelling_alle_zones

def minmaxmatrix (Matrix1, Matrix2, aantal_zones=1425, minmax = "max"):
    Eindmatrix = []
    for i in range (aantal_zones):
        Eindmatrix.append([])
        for j in range (aantal_zones):
            if minmax == "max":
                Eindmatrix[i].append ( max ( Matrix1[i][j], Matrix2[i][j] ) )
            else:
                Eindmatrix[i].append ( min ( Matrix1[i][j], Matrix2[i][j] ) )
    return Eindmatrix
