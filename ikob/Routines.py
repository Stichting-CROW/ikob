import csv
import openpyxl
import xlsxwriter


def matrixvuller ( filenaam,  kolommen, beginrij = 2, beginkolom = 2,   sheetnaam = 'Sheet1') :
    wb2=openpyxl.load_workbook(filenaam)
    sheet2=wb2[sheetnaam]
    matrix = []
    aantal_rijen = sheet2.max_row

    for r in range (beginrij, aantal_rijen + 1):
        matrix.append ([])
        for k in range (beginkolom, kolommen + beginkolom):
            matrixcel = sheet2.cell(row=r, column=k)
            Waarde =  matrixcel.value
            matrix [r-beginrij] .append (Waarde)
    return matrix

def kolomkopvuller (filenaam, sheetnaam) :
    wb2=openpyxl.load_workbook(filenaam)
    sheet2=wb2[sheetnaam]
    kolomkoppen = []

    for k in range ( 2, sheet2.max_column + 1):
            koppencel = sheet2.cell(row=1, column=k)
            Koptekst = koppencel.value
            kolomkoppen .append (Koptekst)
    return kolomkoppen


def zoekrijnummer (i, kolom):
    for j in range ( len ( kolom ) ):
        if kolom[j] == i:
            return j

def zoekSEGkolom(kolomkoppen, tekst):
    for k in range (0,len(kolomkoppen)):
        if kolomkoppen[k] == tekst :
            return k


def lijstvolnullen(lengte) :
    return [0 for _ in range(lengte)]


def transponeren (matrix):
    return [list(i) for i in zip(*matrix)]


def xlswegschrijven (matrix, filenaam, header):
    workbook = xlsxwriter.Workbook ( filenaam + '.xlsx' )
    worksheet = workbook.add_worksheet ( )
    worksheet.write_row (0,0,header)
    for r in range ( 0, len(matrix) ):
        worksheet.write( r+1 , 0 , r+1)
        worksheet.write_row ( r + 1, 1, matrix[r] )
    workbook.close ( )


def xlswegschrijven_totalen (matrix, header, getallenlijst, filenaam, aantal_zones=1425):
    transmatrix = transponeren(matrix)
    workbook = xlsxwriter.Workbook ( filenaam + '.xlsx' )
    worksheet = workbook.add_worksheet ( )
    worksheet.write_row (0,0,header)
    worksheet.write_column (1,0,getallenlijst)
    for r in range ( 0, 1425 ):
        worksheet.write_row ( r + 1, 1, transmatrix[r] )
    workbook.close ( )

def matrixen_maken_voor_Excel_totalen (Totalendirectory, groep):
    vervoerscombis = ['Fiets', 'EFiets', 'Auto', 'OV', 'Fiets_Auto', 'Fiets_OV', 'EFiets_Auto', 'Efiets_OV', 'Auto_OV',
                      'Fiets_Auto_OV', 'Efiets_Auto_OV']
    Matrix = []

    for vvcombis in vervoerscombis:
        if groep is not None:
            Matrix.append(csvlezen (Totalendirectory + vvcombis + '_' + str(groep)))
        else:
            Matrix.append ( csvlezen ( Totalendirectory + vvcombis  ) )
    return Matrix


def getallenlijst_maken (aantal_getallen):
    return list(range(1, aantal_getallen + 1))


def csvlezen (filenaam, aantal_lege_regels=0):
    matrix = []
    filenaam2=filenaam + '.csv'
    with open ( filenaam2, 'r' ) as csvfile:
        reader = csv.reader ( csvfile )
        for i in range (aantal_lege_regels):
            next(reader)
        for row in reader:
            matrix.append ( row )
    with open ( filenaam2, 'r' ) as csvfile:
        reader = csv.reader ( csvfile )
        row_count = sum ( 1 for row in reader )
    uitmatrix = []
    tussenmatrix = []
    if row_count == 1:
        tussenmatrix.append ( matrix[0] )
        for elem in tussenmatrix[0]:
            uitmatrix.append ( float(elem) )
    else:
        for r in range ( 0, row_count - aantal_lege_regels ):
            tussenmatrix.append (matrix[r])
            uitmatrix.append ([])
            for elem in tussenmatrix[r]:
                uitmatrix[r].append ( float (elem) )
    return uitmatrix

def csvintlezen (filenaam, aantal_lege_regels=0):
    matrix = []
    filenaam2=filenaam + '.csv'
    with open ( filenaam2, 'r' ) as csvfile:
        reader = csv.reader ( csvfile )
        for i in range (aantal_lege_regels):
            next(reader)
        for row in reader:
            matrix.append ( row )
    with open ( filenaam2, 'r' ) as csvfile:
        reader = csv.reader ( csvfile )
        row_count = sum ( 1 for row in reader )
    uitmatrix = []
    tussenmatrix = []
    if row_count == 1:
        tussenmatrix.append ( matrix[0] )
        for elem in tussenmatrix[0]:
            uitmatrix.append ( int(elem) )
    else:
        for r in range ( 0, row_count - aantal_lege_regels ):
            tussenmatrix.append (matrix[r])
            uitmatrix.append ([])
            for elem in tussenmatrix[r]:
                uitmatrix[r].append ( int(elem) )
    return uitmatrix

def csvfloatlezen (filenaam, aantal_lege_regels=0):
    matrix = []
    filenaam2=filenaam + '.csv'
    with open ( filenaam2, 'r' ) as csvfile:
        reader = csv.reader ( csvfile )
        for i in range (aantal_lege_regels):
            next(reader)
        for row in reader:
            matrix.append ( row )
    with open ( filenaam2, 'r' ) as csvfile:
        reader = csv.reader ( csvfile )
        row_count = sum ( 1 for row in reader )
    uitmatrix = []
    tussenmatrix = []
    if row_count == 1:
        tussenmatrix.append ( matrix[0] )
        for elem in tussenmatrix[0]:
            uitmatrix.append ( float(elem) )
    else:
        for r in range ( 0, row_count - aantal_lege_regels ):
            tussenmatrix.append (matrix[r])
            uitmatrix.append ([])
            for elem in tussenmatrix[r]:
                uitmatrix[r].append ( float(elem) )
    return uitmatrix

def Omnitrans_csv_inlezen (filenaam, aantal_zones=1425, aantal_lege_regels=4):
    lijst = []
    skimmatrix = []
    filenaam2=filenaam + '.csv'
    with open ( filenaam2, 'r' ) as csvfile:
        reader = csv.reader ( csvfile )
    with open ( filenaam2, 'r' ) as csvfile:
        reader = csv.reader ( csvfile )
        for i in range (aantal_lege_regels):
            next(reader)
        for row in reader:
            lijst.append(row)
    for i in range (aantal_zones):
        skimmatrix.append([])
        for j in range (aantal_zones):
            r = i*1425+j
            skimmatrix[i].append(float(lijst[r][2]))
    return skimmatrix

def csvwegschrijven (matrix, filenaam, soort = "matrix"):
    f = open ( filenaam + '.csv', 'w', newline='' )
    with f:
        writer = csv.writer ( f )
        if soort == "matrix":
            writer.writerows ( matrix )
        else:
            writer.writerow ( matrix )

def csvwegschrijvenmetheader (matrix, filenaam, header, soort = "matrix" ):
    f = open ( filenaam + '.csv', 'w', newline='' )
    with f:
        writer = csv.writer ( f )
        writer.writerow(header)
        if soort == "matrix":
            writer.writerows ( matrix )
        else:
            writer.writerow ( matrix )


def laad_constanten_reistijdvervalscurve (Gewichtendirectory,  vvwijze, groep, motief ):
    zoekkolom = {'werk':2,'winkeldagelijks':6, 'winkelnietdagelijks':10, 'onderwijs':14, 'zorg':18, 'overig':22}
    optelwaarde = {'Auto':0, 'Fiets':1, 'EFiets':2, 'OV':3}
    filenaam2 = Gewichtendirectory + 'Constanten_reistijdvervalscurve.xlsx'
    wb=openpyxl.load_workbook(filenaam2)
    sheet=wb['Alphas']
    sheet2=wb['Omegas']
    alphacel = sheet.cell(row=groep + 3, column= zoekkolom.get(motief) + optelwaarde.get(vvwijze))
    alpha = alphacel.value
    omegacel = sheet2.cell(row=groep + 3, column= zoekkolom.get(motief) + optelwaarde.get(vvwijze))
    omega = omegacel.value
    return alpha, omega

def laad_alle_constanten_reistijdvervalcurve (Gewichtendirectory, motief, aantal_groepen = 20):
    vervoerwijze=['Auto', 'Fiets', 'EFiets', 'OV']

    for vvwijze in vervoerwijze:
        Alphas = []
        Omegas = []
        for groep in range (0, aantal_groepen + 1):
            Alphasnaam = 'Alphas_'+ vvwijze
            Omegasnaam = 'Omegas_' + vvwijze
            Alphaomega = laad_constanten_reistijdvervalscurve(Gewichtendirectory,vvwijze,groep, motief)
            Alphas.append(Alphaomega[0])
            Omegas.append(Alphaomega[1])
        Alphafilenaam = Gewichtendirectory + Alphasnaam
        Omegafilenaam = Gewichtendirectory + Omegasnaam
        csvwegschrijven(Alphas,Alphafilenaam, soort = "lijst")
        csvwegschrijven(Omegas,Omegafilenaam, soort = "lijst")

def maak_totale_buurten_file (Invoerexcel, aantal_zones = 1425):
    Zonelijst_met_buurtsamenstelling = matrixvuller ( Invoerexcel, 1, beginrij=3, beginkolom=1,
                                                               sheetnaam='Excl_studenten' )
    Zonelijst = [sub[0] for sub in Zonelijst_met_buurtsamenstelling]
    Buurtsamenstelling = matrixvuller (Invoerexcel, 20, beginrij=3, beginkolom=2,
                                                 sheetnaam='Excl_studenten' )
    Zonelijstnummer = 0
    Samenstelling_alle_zones = []
    for i in range ( aantal_zones ):
        Samenstelling_alle_zones.append ( [] )
        if i + 1 in Zonelijst:
            for k in range ( 0, 20 ):
                Samenstelling_alle_zones[i].append ( Buurtsamenstelling[Zonelijstnummer][k] )
            Zonelijstnummer = Zonelijstnummer + 1
        else:
            for k in range ( 0, 20 ):
                Samenstelling_alle_zones[i].append ( 0 )
    return Samenstelling_alle_zones

def minmaxmatrix (Matrix1, Matrix2, minmax = "max"):
    Eindmatrix = []
    for i in range (0,len(Matrix1)) :
        Eindmatrix.append([])
        for j in range (0,len(Matrix1)) :
            if minmax == "max":
                Eindmatrix[i].append ( max ( Matrix1[i][j], Matrix2[i][j] ) )
            else:
                Eindmatrix[i].append ( min ( Matrix1[i][j], Matrix2[i][j] ) )
    return Eindmatrix

def minmaxmatrix3(matrix1, matrix2, matrix3, minmax="max"):
    eindmatrix = []
    for i in range(0, len(matrix1)):
        eindmatrix.append([])
        for j in range(0, len(matrix1)):
            if minmax == "max":
                eindmatrix[i].append(max(matrix1[i][j], matrix2[i][j], matrix3[i][j]))
            else:
                eindmatrix[i].append(min(matrix1[i][j], matrix2[i][j], matrix3[i][j]))
    return eindmatrix


def inkomensgroepbepalen(naam):
    if naam[-4:] == 'hoog':
        if naam[-10:] == 'middelhoog':
            return 'middelhoog'
        else:
            return 'hoog'
    elif naam[-4:] == 'laag':
        if naam[-10:] == 'middellaag':
            return 'middellaag'
        else:
            return 'laag'
    else:
        return ''


def vindvoorkeur(naam, mod):
    if 'vk' in naam:
        Beginvk = naam.find ('vk')
        if naam[Beginvk + 2] == "A":
            return 'Auto'
        elif naam[Beginvk + 2] == "N":
            return 'Neutraal'
        elif naam[Beginvk + 2] == "O":
            return 'OV'
        elif naam[Beginvk + 2] == "F":
            return 'Fiets'
        else:
            return ''
    elif 'GratisAuto' in naam:
        if 'GratisAuto_GratisOV' in naam and 'OV' in mod and 'Auto' in mod:
            return 'Neutraal'
        else:
            if 'Auto' in mod:
                return 'Auto'
            else:
                return 'OV'
    elif 'GratisOV' in naam:
        return 'OV'
    else:
        return ''


def enkelegroep(mod, gr) :
    if mod == 'Auto':
        if 'GratisAuto' in gr:
            return 'GratisAuto'
        elif 'Wel' in gr:
            return 'Auto'
        if 'GeenAuto' in gr:
            return 'GeenAuto'
        if 'GeenRijbewijs' in gr:
            return 'GeenRijbewijs'
    if mod == 'OV':
        if 'GratisOV' in gr:
            return 'GratisOV'
        else:
            return 'OV'


def combigroep(mod, gr) :
    string = ''
    if 'Auto' in mod:
        if 'GratisAuto' in gr:
            string = 'GratisAuto'
        elif 'Wel' in gr:
            string = 'Auto'
        if 'GeenAuto' in gr:
            string = 'GeenAuto'
        if 'GeenRijbewijs' in gr:
            string = 'GeenRijbewijs'
    if 'OV' in mod:
        if 'GratisOV' in gr:
            if string == '':
                string = string + 'GratisOV'
            else:
                string = string + '_GratisOV'
        else:
            if string == '':
                string = string + 'OV'
            else:
                string = string + '_OV'
    if 'EFiets' in mod:
        string = string + '_EFiets'
    elif 'Fiets' in mod:
        string = string + '_Fiets'
    return string
